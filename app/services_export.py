"""Geração de relatórios em PDF (reportlab) e Excel (openpyxl).

A ideia é manter as funções **puras**: recebem dicts/listas já consolidados
pelo ``services_financeiro`` e devolvem ``io.BytesIO`` pronto pra
``send_file``. Sem acessar banco aqui.
"""
from io import BytesIO
from datetime import date, datetime

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


MESES_PT = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
            'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #

def _brl(v):
    """Formata número em padrão R$ 1.234,56."""
    if v is None:
        return 'R$ 0,00'
    s = f'{float(v):,.2f}'
    return 'R$ ' + s.replace(',', 'X').replace('.', ',').replace('X', '.')


def _estilos_base(nome_sistema):
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name='TituloSistema', parent=styles['Title'],
        fontSize=14, textColor=colors.HexColor('#2563eb'),
        spaceAfter=2, alignment=0,
    ))
    styles.add(ParagraphStyle(
        name='SubtituloRel', parent=styles['Heading2'],
        fontSize=11, textColor=colors.HexColor('#374151'),
        spaceAfter=12, alignment=0,
    ))
    styles.add(ParagraphStyle(
        name='RodapeMeta', parent=styles['Normal'],
        fontSize=8, textColor=colors.HexColor('#6b7280'),
        spaceAfter=6,
    ))
    return styles


def _cabecalho(story, styles, nome_sistema, titulo_relatorio, subtitulo=None):
    story.append(Paragraph(nome_sistema, styles['TituloSistema']))
    story.append(Paragraph(titulo_relatorio, styles['SubtituloRel']))
    if subtitulo:
        story.append(Paragraph(subtitulo, styles['RodapeMeta']))
    story.append(Paragraph(
        f'Gerado em {datetime.now().strftime("%d/%m/%Y %H:%M")}',
        styles['RodapeMeta']))
    story.append(Spacer(1, 0.4 * cm))


def _estilo_tabela_padrao():
    return TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8.5),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#d1d5db')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1),
         [colors.white, colors.HexColor('#f3f4f6')]),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ])


# --------------------------------------------------------------------------- #
# Fluxo de caixa
# --------------------------------------------------------------------------- #

def fluxo_para_pdf(fluxo, mes, ano, nome_sistema='EducaMais'):
    """``fluxo`` é o dict devolvido por ``services_financeiro.fluxo_caixa``."""
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
    )
    styles = _estilos_base(nome_sistema)
    story = []

    _cabecalho(
        story, styles, nome_sistema,
        'Fluxo de Caixa',
        f'Período: {MESES_PT[mes - 1]} de {ano}',
    )

    # KPIs
    kpi_table = Table([
        ['Entradas', 'Saídas', 'Saldo'],
        [_brl(fluxo['entradas']), _brl(fluxo['saidas']), _brl(fluxo['saldo'])],
    ], colWidths=[5.5 * cm, 5.5 * cm, 5.5 * cm])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f3f4f6')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, 1), 12),
        ('TEXTCOLOR', (0, 1), (0, 1), colors.HexColor('#059669')),
        ('TEXTCOLOR', (1, 1), (1, 1), colors.HexColor('#dc2626')),
        ('TEXTCOLOR', (2, 1), (2, 1),
         colors.HexColor('#2563eb') if fluxo['saldo'] >= 0 else colors.HexColor('#dc2626')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#d1d5db')),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 0.6 * cm))

    # Movimentações
    movs = fluxo.get('movimentacoes', [])
    if movs:
        head = ['Data', 'Descrição', 'Categoria', 'Tipo', 'Valor (R$)']
        rows = [head]
        for m in movs:
            cat = m.categoria.nome if getattr(m, 'categoria', None) else '—'
            sinal = '+' if m.tipo == 'entrada' else '−'
            rows.append([
                m.data.strftime('%d/%m/%Y'),
                Paragraph(m.descricao or '', styles['Normal']),
                cat,
                'Entrada' if m.tipo == 'entrada' else 'Saída',
                f'{sinal} {_brl(m.valor)[3:]}',
            ])
        tbl = Table(rows,
                    colWidths=[2.2 * cm, 7.0 * cm, 3.2 * cm, 2.0 * cm, 3.0 * cm],
                    repeatRows=1)
        tbl.setStyle(_estilo_tabela_padrao())
        story.append(tbl)
    else:
        story.append(Paragraph('Nenhuma movimentação no período.', styles['Normal']))

    doc.build(story)
    buf.seek(0)
    return buf


def fluxo_para_xlsx(fluxo, mes, ano, nome_sistema='EducaMais'):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Movimentações'

    bold = Font(bold=True, color='FFFFFF')
    fill = PatternFill('solid', fgColor='2563eb')
    center = Alignment(horizontal='center', vertical='center')
    border_thin = Border(*(Side(style='thin', color='D1D5DB'),) * 4)

    headers = ['Data', 'Descrição', 'Categoria', 'Tipo', 'Valor (R$)']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = bold
        c.fill = fill
        c.alignment = center

    for r, m in enumerate(fluxo.get('movimentacoes', []), 2):
        ws.cell(row=r, column=1, value=m.data).number_format = 'DD/MM/YYYY'
        ws.cell(row=r, column=2, value=m.descricao)
        ws.cell(row=r, column=3,
                value=m.categoria.nome if getattr(m, 'categoria', None) else '')
        ws.cell(row=r, column=4, value='Entrada' if m.tipo == 'entrada' else 'Saída')
        sinal = 1 if m.tipo == 'entrada' else -1
        v = ws.cell(row=r, column=5, value=float(m.valor) * sinal)
        v.number_format = 'R$ #,##0.00;[Red]-R$ #,##0.00'

    widths = [12, 50, 22, 12, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                            min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = border_thin

    # Aba Resumo
    ws2 = wb.create_sheet('Resumo')
    ws2['A1'] = nome_sistema
    ws2['A1'].font = Font(bold=True, size=14, color='2563eb')
    ws2['A2'] = 'Fluxo de Caixa'
    ws2['A2'].font = Font(bold=True, size=12)
    ws2['A3'] = f'{MESES_PT[mes - 1]} / {ano}'
    ws2['A4'] = f'Gerado em {datetime.now().strftime("%d/%m/%Y %H:%M")}'

    ws2['A6'] = 'Entradas'
    ws2['A7'] = 'Saídas'
    ws2['A8'] = 'Saldo'
    ws2['B6'] = float(fluxo['entradas'])
    ws2['B7'] = float(fluxo['saidas'])
    ws2['B8'] = float(fluxo['saldo'])
    for r in (6, 7, 8):
        ws2.cell(row=r, column=2).number_format = 'R$ #,##0.00;[Red]-R$ #,##0.00'
        ws2.cell(row=r, column=1).font = Font(bold=True)
    ws2.column_dimensions['A'].width = 18
    ws2.column_dimensions['B'].width = 18

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# --------------------------------------------------------------------------- #
# Mensalidades
# --------------------------------------------------------------------------- #

def _status_mensalidade(m):
    """Retorna ('Paga' | 'Aberta' | 'Vencida' | 'Sem boleto', cor)."""
    hoje = date.today()
    boleto_pago = any(b.status == 'pago' for b in (m.boletos or []))
    if boleto_pago:
        return 'Paga', '#059669'
    if not m.boletos:
        return 'Sem boleto', '#6b7280'
    if m.vencimento and m.vencimento < hoje:
        return 'Vencida', '#dc2626'
    return 'Aberta', '#2563eb'


def mensalidades_para_pdf(mensalidades, mes, ano, nome_sistema='EducaMais'):
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
    )
    styles = _estilos_base(nome_sistema)
    story = []
    _cabecalho(
        story, styles, nome_sistema, 'Mensalidades',
        f'Período: {MESES_PT[mes - 1]} de {ano}'
            f' • {len(mensalidades)} aluno(s)',
    )

    if not mensalidades:
        story.append(Paragraph('Nenhuma mensalidade no período.', styles['Normal']))
    else:
        head = ['Aluno', 'Responsável', 'Vencimento', 'Status', 'Valor']
        rows = [head]
        total = 0
        recebido = 0
        for m in mensalidades:
            status, _cor = _status_mensalidade(m)
            resp = m.responsavel.nome if m.responsavel else 'O próprio aluno'
            rows.append([
                Paragraph(m.aluno.nome, styles['Normal']),
                Paragraph(resp, styles['Normal']),
                m.vencimento.strftime('%d/%m/%Y') if m.vencimento else '—',
                status,
                _brl(m.valor),
            ])
            total += float(m.valor or 0)
            if status == 'Paga':
                recebido += float(m.valor or 0)

        rows.append(['', '', '', 'Total previsto:', _brl(total)])
        rows.append(['', '', '', 'Total recebido:', _brl(recebido)])

        tbl = Table(rows,
                    colWidths=[5.5 * cm, 4.8 * cm, 2.4 * cm, 2.4 * cm, 2.4 * cm],
                    repeatRows=1)
        st = _estilo_tabela_padrao()
        n = len(rows)
        st.add('FONTNAME', (0, n - 2), (-1, n - 1), 'Helvetica-Bold')
        st.add('BACKGROUND', (0, n - 2), (-1, n - 1),
               colors.HexColor('#f3f4f6'))
        tbl.setStyle(st)
        story.append(tbl)

    doc.build(story)
    buf.seek(0)
    return buf


def mensalidades_para_xlsx(mensalidades, mes, ano, nome_sistema='EducaMais'):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Mensalidades'

    headers = ['Aluno', 'Responsável', 'Mês/Ano', 'Vencimento', 'Status',
               'Valor (R$)', 'Boletos']
    bold = Font(bold=True, color='FFFFFF')
    fill = PatternFill('solid', fgColor='2563eb')

    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = bold
        c.fill = fill

    for r, m in enumerate(mensalidades, 2):
        status, _ = _status_mensalidade(m)
        ws.cell(row=r, column=1, value=m.aluno.nome)
        ws.cell(row=r, column=2,
                value=m.responsavel.nome if m.responsavel else 'O próprio aluno')
        ws.cell(row=r, column=3, value=f'{m.mes:02d}/{m.ano}')
        ws.cell(row=r, column=4, value=m.vencimento).number_format = 'DD/MM/YYYY'
        ws.cell(row=r, column=5, value=status)
        ws.cell(row=r, column=6,
                value=float(m.valor or 0)).number_format = 'R$ #,##0.00'
        ws.cell(row=r, column=7, value=len(m.boletos or []))

    widths = [28, 28, 12, 12, 12, 14, 9]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# --------------------------------------------------------------------------- #
# Inadimplentes
# --------------------------------------------------------------------------- #

def inadimplentes_para_pdf(lista, mes_nome, ano, nome_sistema='EducaMais',
                           escopo='mes'):
    """``lista`` é o retorno de ``services_financeiro.inadimplentes``:

    cada item é ``{aluno, responsavel, boletos: [...], total_devido}``.
    """
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
    )
    styles = _estilos_base(nome_sistema)
    story = []

    titulo_periodo = (f'Inadimplentes de {mes_nome}/{ano}' if escopo == 'mes'
                      else 'Inadimplentes — Todos os períodos em atraso')
    _cabecalho(story, styles, nome_sistema, 'Relatório de Inadimplência',
               titulo_periodo)

    if not lista:
        story.append(Paragraph('Nenhum inadimplente no período.',
                               styles['Normal']))
    else:
        head = ['Aluno', 'Responsável / Contato', 'Boletos em atraso', 'Total']
        rows = [head]
        total_geral = 0
        hoje = date.today()
        for it in lista:
            aluno = it['aluno']
            resp = it.get('responsavel')
            boletos = it.get('boletos', [])

            tel = (resp.telefone if resp else getattr(aluno, 'telefone', None)) or 'sem telefone'
            nome_pagador = resp.nome if resp else 'O próprio aluno'
            contato = f'{nome_pagador}\n{tel}'

            linhas_boletos = []
            for b in boletos:
                dias = (hoje - b.vencimento).days if b.vencimento else 0
                linhas_boletos.append(
                    f'#{b.id} • venc. {b.vencimento.strftime("%d/%m/%Y")} • '
                    f'{dias}d • {_brl(b.valor)}'
                )

            rows.append([
                Paragraph(aluno.nome, styles['Normal']),
                Paragraph(contato.replace('\n', '<br/>'), styles['Normal']),
                Paragraph('<br/>'.join(linhas_boletos), styles['Normal']),
                _brl(it['total_devido']),
            ])
            total_geral += float(it['total_devido'] or 0)

        rows.append(['', '', 'Total geral em atraso:', _brl(total_geral)])

        tbl = Table(rows,
                    colWidths=[4.5 * cm, 5.0 * cm, 6.5 * cm, 2.5 * cm],
                    repeatRows=1)
        st = _estilo_tabela_padrao()
        st.add('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold')
        st.add('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#fef2f2'))
        tbl.setStyle(st)
        story.append(tbl)

    doc.build(story)
    buf.seek(0)
    return buf


# --------------------------------------------------------------------------- #
# Boletos
# --------------------------------------------------------------------------- #

def boletos_para_xlsx(boletos, mes, ano, nome_sistema='EducaMais'):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Boletos'

    headers = ['ID', 'Aluno', 'Responsável', 'Tipo', 'Status',
               'Vencimento', 'Pago em', 'Valor (R$)', 'Cora ID']
    bold = Font(bold=True, color='FFFFFF')
    fill = PatternFill('solid', fgColor='2563eb')
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = bold
        c.fill = fill

    for r, b in enumerate(boletos, 2):
        mensa = b.mensalidade
        aluno_nome = mensa.aluno.nome if (mensa and mensa.aluno) else '—'
        resp_nome = (mensa.responsavel.nome if (mensa and mensa.responsavel)
                     else 'O próprio aluno')
        ws.cell(row=r, column=1, value=b.id)
        ws.cell(row=r, column=2, value=aluno_nome)
        ws.cell(row=r, column=3, value=resp_nome)
        ws.cell(row=r, column=4, value=b.tipo_cobranca or 'cora')
        ws.cell(row=r, column=5, value=b.status)
        ws.cell(row=r, column=6, value=b.vencimento).number_format = 'DD/MM/YYYY'
        if b.pago_em:
            ws.cell(row=r, column=7,
                    value=b.pago_em).number_format = 'DD/MM/YYYY'
        ws.cell(row=r, column=8,
                value=float(b.valor or 0)).number_format = 'R$ #,##0.00'
        ws.cell(row=r, column=9, value=b.cora_boleto_id or '')

    widths = [6, 28, 24, 12, 10, 12, 12, 14, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# --------------------------------------------------------------------------- #
# Relatório de status dos alunos
# --------------------------------------------------------------------------- #

def relatorio_status_pdf(snapshot, filtro_turma=None,
                         nome_instituicao='EducaMais'):
    """Gera PDF com KPIs e tabelas a partir do snapshot de
    ``services_relatorios.snapshot_completo``.

    Sem gráficos por enquanto (decisão de Fase 4) — só KPIs e tabelas.
    """
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
    )
    styles = _estilos_base(nome_instituicao)
    story = []

    subtitulo = 'Recorte: ' + (filtro_turma if filtro_turma else 'todas as turmas')
    _cabecalho(
        story, styles, nome_instituicao,
        'Relatório de Alunos · Status & Histórico',
        subtitulo,
    )

    k = snapshot['kpis']

    # ── KPIs ────────────────────────────────────────────────────────────
    kpi_head = ['Ativos', 'Formados', 'Evadidos', 'Transferidos', 'Taxa de evasão']
    kpi_vals = [str(k['ativos']), str(k['formados']),
                str(k['evadidos']), str(k['transferidos']),
                f"{k['taxa_evasao']}%"]
    kpi_table = Table([kpi_head, kpi_vals],
                      colWidths=[3.4 * cm, 3.4 * cm, 3.4 * cm, 3.4 * cm, 3.4 * cm])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f3f4f6')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, 1), 14),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0, 1), (0, 1), colors.HexColor('#059669')),
        ('TEXTCOLOR', (1, 1), (1, 1), colors.HexColor('#2563eb')),
        ('TEXTCOLOR', (2, 1), (2, 1), colors.HexColor('#d97706')),
        ('TEXTCOLOR', (3, 1), (3, 1), colors.HexColor('#64748b')),
        ('TEXTCOLOR', (4, 1), (4, 1), colors.HexColor('#dc2626')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#d1d5db')),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 0.3 * cm))
    story.append(Paragraph(
        f"Total cadastrado: <b>{k['total']}</b> aluno(s)"
        + (f" · {k['sem_vinculo']} sem vínculo de turma" if k['sem_vinculo'] else ''),
        styles['RodapeMeta']))
    story.append(Spacer(1, 0.6 * cm))

    # ── Detalhamento por turma ──────────────────────────────────────────
    story.append(Paragraph('Distribuição por turma', styles['SubtituloRel']))
    por_turma = snapshot.get('por_turma') or []
    if por_turma:
        rows = [['Turma', 'Ativos', 'Formados', 'Evadidos', 'Transf.', 'Total']]
        for r in por_turma:
            rows.append([
                Paragraph(r['turma'].nome, styles['Normal']),
                str(r['ativos']),
                str(r['formados']),
                str(r['evadidos']),
                str(r['transferidos']),
                str(r['total']),
            ])
        tbl = Table(
            rows,
            colWidths=[7.0 * cm, 2.0 * cm, 2.0 * cm, 2.0 * cm, 2.0 * cm, 2.0 * cm],
            repeatRows=1,
        )
        estilo = _estilo_tabela_padrao()
        estilo.add('ALIGN', (1, 0), (-1, -1), 'CENTER')
        tbl.setStyle(estilo)
        story.append(tbl)
    else:
        story.append(Paragraph('Nenhuma turma com matrículas no recorte.',
                               styles['Normal']))

    story.append(Spacer(1, 0.6 * cm))

    # ── Histórico anual ─────────────────────────────────────────────────
    story.append(Paragraph('Saídas por ano', styles['SubtituloRel']))
    hist = snapshot.get('historico') or []
    if hist:
        rows = [['Ano', 'Formados', 'Evadidos', 'Transferidos', 'Total']]
        for h in hist:
            total = h['formados'] + h['evadidos'] + h['transferidos']
            rows.append([
                str(h['ano']),
                str(h['formados']),
                str(h['evadidos']),
                str(h['transferidos']),
                str(total),
            ])
        tbl = Table(
            rows,
            colWidths=[3.0 * cm, 3.0 * cm, 3.0 * cm, 3.5 * cm, 3.0 * cm],
            repeatRows=1,
        )
        estilo = _estilo_tabela_padrao()
        estilo.add('ALIGN', (0, 0), (-1, -1), 'CENTER')
        tbl.setStyle(estilo)
        story.append(tbl)
    else:
        story.append(Paragraph('Ainda não há saídas registradas.',
                               styles['Normal']))

    doc.build(story)
    buf.seek(0)
    return buf
